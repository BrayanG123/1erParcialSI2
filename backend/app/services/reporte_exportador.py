"""
reporte_exportador.py — Convierte el resultado de un reporte QBE en
archivos descargables: Excel (.xlsx) y PDF.

Recibe el dict que devuelve ejecutar_qbe() (entidad, tipo_reporte, total,
columnas, filas) y produce los BYTES del archivo — listos para:
  - devolverse como descarga (endpoint /reportes/exportar)
  - adjuntarse a un correo (endpoint /reportes/enviar-correo)

Librerías:
  - openpyxl → Excel (estándar de facto en Python)
  - fpdf2    → PDF (pura Python, sin dependencias del sistema)
"""
import io
from datetime import datetime
from enum import Enum


# ─────────────────────────────────────────────────────────────
# Utilidades
# ─────────────────────────────────────────────────────────────

def _celda_a_texto(valor) -> str:
    """
    Normaliza cualquier valor de la BD a texto presentable:
    enums → su valor, fechas → dd/mm/aaaa HH:MM, floats → 2 decimales.
    """
    if valor is None:
        return ""
    if isinstance(valor, Enum):
        return str(valor.value)
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")
    if isinstance(valor, float):
        return f"{valor:.2f}"
    return str(valor)


def _titulo_reporte(resultado: dict) -> str:
    tipo = "Ejecutivo" if resultado["tipo_reporte"] == "agrupado" else "Gerencial"
    return f"Reporte {tipo} — {resultado['entidad'].capitalize()}"


# ─────────────────────────────────────────────────────────────
# EXCEL (.xlsx)
# ─────────────────────────────────────────────────────────────

def generar_excel(resultado: dict) -> bytes:
    """Genera un .xlsx con encabezado estilizado y columnas auto-ajustadas."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hoja = wb.active
    hoja.title = "Reporte"

    columnas = resultado["columnas"]
    filas = resultado["filas"]

    # ── Fila 1: título del reporte ──
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columnas), 1))
    celda_titulo = hoja.cell(row=1, column=1, value=_titulo_reporte(resultado))
    celda_titulo.font = Font(bold=True, size=14, color="4F46E5")

    # ── Fila 2: metadatos ──
    hoja.cell(
        row=2, column=1,
        value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
              f"Total de registros: {resultado['total']}",
    ).font = Font(size=9, color="888888")

    # ── Fila 4: encabezados de columna ──
    fila_encabezado = 4
    relleno = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for col_idx, nombre in enumerate(columnas, start=1):
        celda = hoja.cell(row=fila_encabezado, column=col_idx, value=nombre)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center")

    # ── Datos ──
    for fila_idx, fila in enumerate(filas, start=fila_encabezado + 1):
        for col_idx, nombre in enumerate(columnas, start=1):
            valor = fila.get(nombre)
            # Números van como números (para que Excel pueda sumarlos);
            # el resto como texto normalizado
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                hoja.cell(row=fila_idx, column=col_idx, value=valor)
            else:
                hoja.cell(row=fila_idx, column=col_idx, value=_celda_a_texto(valor))

    # ── Ancho de columnas: según el contenido más largo (tope 45) ──
    for col_idx, nombre in enumerate(columnas, start=1):
        ancho = max(
            len(str(nombre)),
            *(len(_celda_a_texto(f.get(nombre))) for f in filas[:200])
        ) if filas else len(str(nombre))
        hoja.column_dimensions[get_column_letter(col_idx)].width = min(ancho + 3, 45)

    # Congelar el encabezado (al hacer scroll, los títulos quedan fijos)
    hoja.freeze_panes = hoja.cell(row=fila_encabezado + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────────

def _latin1(texto: str, maximo: int = 60) -> str:
    """
    Las fuentes estándar de PDF solo soportan latin-1 (incluye á é í ó ú ñ).
    Caracteres fuera de ese set (emojis, etc.) se reemplazan por '?'.
    También recorta el texto para que quepa en la celda.
    """
    t = texto.encode("latin-1", "replace").decode("latin-1")
    if len(t) > maximo:
        return t[: maximo - 3] + "..."
    return t


def generar_pdf(resultado: dict) -> bytes:
    """Genera un PDF apaisado con la tabla del reporte."""
    from fpdf import FPDF

    columnas = resultado["columnas"]
    filas = resultado["filas"]

    # Apaisado (L) para que entren más columnas
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    # ── Título y metadatos ──
    pdf.set_font("helvetica", "B", 14)
    pdf.set_text_color(79, 70, 229)  # índigo, como la web
    pdf.cell(0, 8, _latin1(_titulo_reporte(resultado), 100), new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("helvetica", "", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(
        0, 5,
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
        f"Total de registros: {resultado['total']}",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.ln(3)

    if not filas:
        pdf.set_font("helvetica", "I", 11)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 10, "Sin resultados para los criterios indicados.")
        return bytes(pdf.output())

    # ── Anchos de columna proporcionales al contenido ──
    ancho_util = pdf.w - pdf.l_margin - pdf.r_margin
    largos = []
    for col in columnas:
        max_len = max(len(str(col)), *(len(_celda_a_texto(f.get(col))) for f in filas[:100]))
        largos.append(min(max_len, 40))   # tope para que una columna no devore todo
    total_largos = sum(largos) or 1
    anchos = [max(18.0, ancho_util * l / total_largos) for l in largos]
    # Reescalar si el mínimo de 18mm nos pasó del ancho disponible
    factor = ancho_util / sum(anchos)
    anchos = [a * factor for a in anchos]

    alto_fila = 7

    def dibujar_encabezado():
        pdf.set_font("helvetica", "B", 8)
        pdf.set_fill_color(79, 70, 229)
        pdf.set_text_color(255, 255, 255)
        for ancho, col in zip(anchos, columnas):
            pdf.cell(ancho, alto_fila, _latin1(str(col), 30), border=1, fill=True)
        pdf.ln(alto_fila)

    dibujar_encabezado()

    # ── Filas (con cebra y repetición de encabezado al saltar de página) ──
    pdf.set_text_color(40, 40, 40)
    for i, fila in enumerate(filas):
        if pdf.get_y() + alto_fila > pdf.h - 12:   # ¿no entra otra fila?
            pdf.add_page()
            dibujar_encabezado()
            pdf.set_text_color(40, 40, 40)

        pdf.set_font("helvetica", "", 8)
        pdf.set_fill_color(243, 244, 246)          # gris clarito para filas pares
        for ancho, col in zip(anchos, columnas):
            # caracteres por celda según su ancho (≈ 0.55 mm por carácter a 8pt)
            max_chars = max(8, int(ancho / 1.6))
            pdf.cell(
                ancho, alto_fila,
                _latin1(_celda_a_texto(fila.get(col)), max_chars),
                border=1, fill=(i % 2 == 1),
            )
        pdf.ln(alto_fila)

    return bytes(pdf.output())
